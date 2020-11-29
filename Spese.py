from openpyxl import load_workbook
from datetime import date

from functools import partial

from openpyxl.styles import Alignment

import tkinter as tk
import tkinter.ttk as ttk

treaker = load_workbook("Expensive_treaker.xlsx")
giornoConfermato = "01"
meseConfermato = "01"


mesi = {
    "01": treaker["Gennaio"],
    "02": treaker["Febbraio"],
    "03": treaker["Marzo"],
    "04": treaker["Aprile"],
    "05": treaker["Maggio"],
    "06": treaker["Giugno"],
    "07": treaker["Luglio"],
    "08": treaker["Agosto"],
    "09": treaker["Settembre"],
    "10": treaker["Ottobre"],
    "11": treaker["Novembre"],
    "12": treaker["Dicembre"],
    "13": treaker["Riepilogo"]
}

mesiStrInt = {
    1: "Gennaio",
    2: "Febbraio",
    3: "Marzo",
    4: "Aprile",
    5: "Maggio",
    6: "Giugno",
    7: "Luglio",
    8: "Agosto",
    9: "Settembre",
    10: "Ottobre",
    11: "Novembre",
    12: "Dicembre"
}

mesiNum = {
    "Gennaio": "01",
    "Febbraio": "02",
    "Marzo": "03",
    "Aprile": "04",
    "Maggio": "05",
    "Giugno": "06",
    "Luglio": "07",
    "Agosto": "08",
    "Settembre": "09",
    "Ottobre": "10",
    "Novembre": "11",
    "Dicembre": "12"
}

#---------------------------------LOGICA-----------------------------#

    #-----LOGICA BASE------#

def total_update(foglio):
    totale = 0
    x = 2
    while foglio.cell(x, 1).value != None:
        totale += int(foglio.cell(x, 3).value)
        x += 1
    foglio.cell(2, 5).value = totale
    mese = int(mesiNum[foglio.title])
    treaker["Riepilogo"].cell(mese+1, 2).value = totale
    treaker["Riepilogo"].cell(2, 5).value = 0
    for i in range(2,13):
        treaker["Riepilogo"].cell(2,5).value += int(treaker["Riepilogo"].cell(i,2).value)

def getFoglio(data):
    mese = data[5:7]
    return mesi[mese]

def aggiornaExcel(data, nome, valore):
    print(data)
    foglio = getFoglio(data)

    x = 2
    while foglio.cell(x, 1).value is not None:
        x += 1

    foglio.cell(x, 1).value = data
    foglio.cell(x, 1).alignment = Alignment(horizontal='center')
    foglio.cell(x, 2).value = nome
    foglio.cell(x, 2).alignment = Alignment(horizontal='center')
    foglio.cell(x, 3).value = -int(valore)
    foglio.cell(x, 3).alignment = Alignment(horizontal='center')
    foglio.cell(x, 3).number_format = '#,##0.00€'

    total_update(foglio)
    treaker.save("Expensive_treaker.xlsx")

    #------LOGICA GRAFICA--------#

def apri_inserisci():
    f_inserisci.lift()

def apri_visualizza():
    f_visualizza.lift()

def conferma_data():
    print("confermata")
    f_inserisci_pagamento.lift()

giorno_selezionato = 1
mese_selezionato = 1

def converti_giorno_mese_string(data):
    data = str(data)
    if(len(str(data)) < 2):
        data = "0" + str(data)
    return data

def aggiorna_label():
    global giorno_selezionato
    global mese_selezionato
    l_inserisci_data_info.config(text=f"Data Selezionata: {giorno_selezionato}/{mese_selezionato}/2019")

def giorno_selezionato_f(par):
    global giorno_selezionato
    if len(par.widget.curselection())==0:
        return
    giorno_selezionato = par.widget.curselection()[0]+1
    aggiorna_label()

def mese_selezionato_f(par):
    global mese_selezionato
    if len(par.widget.curselection())==0:
        return
    mese_selezionato = par.widget.curselection()[0]+1
    aggiorna_label()

def seleziona_oggi():
    global giorno_selezionato
    global mese_selezionato
    data = date.today()
    oggi = data.strftime("%d")
    mese = data.strftime("%m")
    giorno_selezionato = int(oggi)
    mese_selezionato = int(mese)

    aggiorna_label()

def conferma_pagamento():
    print(e_inserisci_pagamento_nome.get())
    print(e_inserisci_pagamento_valore.get())
    aggiornaExcel(f"2019/{converti_giorno_mese_string(mese_selezionato)}/{converti_giorno_mese_string(giorno_selezionato)}", e_inserisci_pagamento_nome.get(), e_inserisci_pagamento_valore.get())

def visualizza_selezionato_f(par):
    if len(par.widget.curselection())==0:
        return
    selezionato = converti_giorno_mese_string(par.widget.curselection()[0]+1)
    foglio = mesi[selezionato]
    valore = foglio.cell(2, 5).value

    if valore is None:
        valore = "0"

    valore = str(valore) + "€"

    l_visualizza_risultato.config(text=f"Pagato: {valore}")

#------------------DICHIARAZIONE----------------------#

root = tk.Tk()

    #-------FRAME MAIN--------------#

f_main = tk.Frame(root)

b_main_inserisci = tk.Button(f_main, text="Inserisci", command=apri_inserisci)
b_main_visualizza = tk.Button(f_main, text="Visualizza", command=apri_visualizza)

    #-------FRAME LAVORO------------#

f_lavoro = tk.Frame(root)

    #-------FRAME INSERIMENTO-------#
f_inserisci = tk.Frame(f_lavoro)

l_inserisci_giorno_mese = tk.Label(f_inserisci, text="Seleziona giorno e mese:")

f_inserisci_giorno_mese = tk.Frame(f_inserisci)
lb_inserisci_giorno = tk.Listbox(f_inserisci_giorno_mese, height=12)
lb_inserisci_mese = tk.Listbox(f_inserisci_giorno_mese, height=12)
b_inserisci_oggi = tk.Button(f_inserisci, text="Seleziona OGGI", command=seleziona_oggi)
l_inserisci_data_info = tk.Label(f_inserisci, text="Data Inserita: 01/01/2019")
b_inserisci_conferma = tk.Button(f_inserisci, text="Conferma Data", command=conferma_data)

f_inserisci_pagamento = tk.Frame(f_lavoro)
l_inserisci_pagamento_nome = tk.Label(f_inserisci_pagamento, text="Inserire nome pagamento:")
e_inserisci_pagamento_nome = tk.Entry(f_inserisci_pagamento)

l_inserisci_pagamento_valore = tk.Label(f_inserisci_pagamento, text="Inserire valore pagato:")
e_inserisci_pagamento_valore = tk.Entry(f_inserisci_pagamento)

b_inserisci_pagamento_conferma = tk.Button(f_inserisci_pagamento, text="Conferma Pagamento", command=conferma_pagamento)

    #-------FRAME VISUALIZZAZIONE-------#
f_visualizza = tk.Frame(f_lavoro)
lb_visualizza_lista = tk.Listbox(f_visualizza, height=13)

l_visualizza_risultato = tk.Label(f_visualizza, text="Qui il valore")

#---------------------------------POPOLAMENTO------------------------#

for i in range(1, 32):
    lb_inserisci_giorno.insert(tk.END, str(i) if i >= 10 else ("0" + str(i)))
lb_inserisci_giorno.bind("<<ListboxSelect>>", giorno_selezionato_f)

for i in range(1, 13):
    #lb_inserisci_mese.insert(tk.END, str(i) if i >= 10 else ("0" + str(i)))
    lb_inserisci_mese.insert(tk.END, mesiStrInt[i])
lb_inserisci_mese.bind("<<ListboxSelect>>", mese_selezionato_f)

for i in range(1, 13):
    lb_visualizza_lista.insert(tk.END, mesiStrInt[i])
lb_visualizza_lista.insert(tk.END, "TOTALE")
lb_visualizza_lista.bind("<<ListboxSelect>>", visualizza_selezionato_f)

#---------------------------------PACKING----------------------------#

b_main_inserisci.pack(side="left", fill="both", expand=False)
b_main_visualizza.pack(side="right", fill="both", expand=False)

f_main.pack(side="top", fill="both", expand=False)

f_lavoro.pack(side="top", fill="both", expand=True)

l_inserisci_giorno_mese.pack(side="top", fill="both", expand=False)

lb_inserisci_giorno.pack(side="left", fill="both", expand=True)
lb_inserisci_mese.pack(side="right", fill="both", expand=True)

f_inserisci_giorno_mese.pack(side="top", fill="both", expand=False)

b_inserisci_oggi.pack(side="top", fill="both", padx= 10, expand=False)
b_inserisci_conferma.pack(side="bottom", fill="both", padx=10, expand=False)
l_inserisci_data_info.pack(side="bottom", fill="both", padx=50, pady=50, expand=True)



#f_inserisci.pack(side="top", fill="both", expand=True)
f_inserisci.place(x=0,y=0, relwidth=1, relheight=1)

l_inserisci_pagamento_nome.pack(side="top", fill="both", expand=False)
e_inserisci_pagamento_nome.pack(side="top", fill="both", expand=False)

l_inserisci_pagamento_valore.pack(side="top", fill="both", expand=False)
e_inserisci_pagamento_valore.pack(side="top", fill="both", expand=False)

b_inserisci_pagamento_conferma.pack(side="top", fill="both", expand=False)

#f_inserisci_pagamento.pack(side="top", fill="both", expand=True)
f_inserisci_pagamento.place(x=0,y=0, relwidth=1, relheight=1)

lb_visualizza_lista.pack(side="top", fill="both", expand=False)

l_visualizza_risultato.pack(side="top", fill="both",expand=True)

f_visualizza.place(x=0,y=0, relwidth=1, relheight=1)

#---------------------------------MAIN-------------------------------#

seleziona_oggi()

f_inserisci.lift()

root.wm_geometry("600x600")
root.mainloop()